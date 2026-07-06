"""Deterministic parser for the campaign metadata sheet (``.xlsx`` / ``.csv``).

This is the first Phase-3 extractor seamed by ``extract/__init__.py`` and designed
in ``docs/extraction.md`` (§2 evidence rules, §3 locators, §5/§6 field map). It is
purely deterministic: it emits *only* values literally present in the sheet, each
carrying a single ``spreadsheet`` evidence entry captured at read time. Missing
fields are simply not emitted — the Slice-3B draft builder decides ``missing``.
Nothing here guesses, and nothing reaches a network or a model.

Public entry points:
  - :func:`parse_structured` — the mapped scalar ``ExtractedField`` list.
  - :func:`parse_contributors` — experimenters, which land in ``attribution`` (3B),
    NOT in ``fields`` — kept out of :class:`ExtractedField` on purpose.
  - :data:`FIELD_MAP` — sheet ``field`` name → ``(official dotted path, py_type)``.

Type coercion is conservative: a value is coerced to ``float``/``int`` ONLY when
:data:`FIELD_MAP` says so (temperature, mass fractions, pellet diameter, n_scans).
Strings such as beamline ``"15-2"`` or formula ``"CuO2"`` are never coerced. If a
declared numeric coercion fails, the raw string is kept and the field is marked
``needs_confirmation`` rather than crashing or guessing.
"""

from __future__ import annotations

import csv
from collections import namedtuple
from pathlib import Path
from typing import Callable

from . import ExtractedField


def _to_number(raw):
    """Coerce to the tightest JSON number: ``int`` when the text parses as an
    integer, else ``float``. Both serialize as JSON ``number`` — e.g. ``298``
    stays an int and ``298.5`` becomes a float — so a physically-fractional value
    (a temperature) may be whole or fractional without guessing. Raises
    ``ValueError`` if neither parses; that is caught by :func:`_coerce`, which then
    keeps the raw string as ``needs_confirmation`` (never a fabricated number).
    """
    text = str(raw).strip()
    try:
        return int(text)
    except ValueError:
        return float(text)


# ---------------------------------------------------------------------------
# Sheet field name -> (official dotted JSON-path, coercer).
# Built from docs/extraction.md §5 (value -> draft field) and §6 (draft key ==
# official path, verified against schema/isaac_record_v1.json). The coercer is one
# of {str, float, int, _to_number}; ``str`` means "keep verbatim, never coerce",
# and ``_to_number`` accepts int OR float (both JSON numbers) for a value that may
# legitimately be whole or fractional.
# ---------------------------------------------------------------------------
FIELD_MAP: dict[str, tuple[str, Callable]] = {
    # Campaign Info -> system.facility.* / system.technique
    "facility_name": ("system.facility.facility_name", str),
    "organization": ("system.facility.organization", str),
    "site": ("system.facility.site", str),
    "beamline": ("system.facility.beamline", str),
    "endstation": ("system.facility.endstation", str),
    "technique": ("system.technique", str),
    # Campaign Info -> system.configuration.* (designated open namespace)
    "proposal_id": ("system.configuration.proposal_id", str),
    "session_id": ("system.configuration.session_id", str),
    # Campaign Info -> timestamps.*
    "acquired_start_utc": ("timestamps.acquired_start_utc", str),
    "acquired_end_utc": ("timestamps.acquired_end_utc", str),
    "created_utc": ("timestamps.created_utc", str),
    # Sample -> sample.material.* / sample.sample_form
    "material_name": ("sample.material.name", str),
    "formula": ("sample.material.formula", str),
    "provenance": ("sample.material.provenance", str),
    "sample_form": ("sample.sample_form", str),
    # Sample -> sample.composition.* (open object) / sample.geometry.* (open)
    "CuO2_mass_fraction": ("sample.composition.CuO2_mass_fraction", float),
    "sucrose_mass_fraction": ("sample.composition.sucrose_mass_fraction", float),
    "pellet_diameter_mm": ("sample.geometry.pellet_diameter_mm", float),
    # Configurations -> context.*
    "environment": ("context.environment", str),
    "temperature_K": ("context.temperature_K", _to_number),
    "atmosphere": ("context.thermodynamics.atmosphere", str),
    # Configurations -> system.configuration.* (open namespace)
    "monochromator_crystal": ("system.configuration.monochromator_crystal", str),
    "spectrometer_geometry": ("system.configuration.spectrometer_geometry", str),
    "detector_model": ("system.configuration.detector_model", str),
    "n_scans": ("system.configuration.n_scans", int),
}

# Experimenter rows do NOT map to a ``fields`` path; they belong in
# ``attribution.contributors[]`` (docs/extraction.md §5), wired up in Slice-3B.
# They are surfaced only through :func:`parse_contributors`, never as
# :class:`ExtractedField`.
CONTRIBUTOR_FIELDS: tuple[str, ...] = ("lead_experimenter", "co_experimenter")

# Fields present in the synthetic sheet but intentionally NOT in FIELD_MAP:
# incident_energy_start_eV / incident_energy_end_eV -> measurement.series (3B),
# qc_status -> measurement.qc (3B). They are skipped here, never guessed into a
# scalar field.

# A normalized sheet row shared by the CSV and XLSX readers.
_Row = namedtuple("_Row", "section field value unit locator")


def _evidence(source_file: str, locator: str, quote: str | None) -> dict:
    """One ``spreadsheet`` evidence entry (models.evidence shape; drops Nones)."""
    entry = {
        "source_type": "spreadsheet",
        "source_file": source_file,
        "locator": locator,
        "quote": quote,
    }
    return {k: v for k, v in entry.items() if v is not None}


def _coerce(raw, py_type: Callable):
    """Coerce ``raw`` per the declared coercer. Returns ``(value, status)``.

    ``str`` passes through verbatim (status ``verified``). A numeric coercer
    (``int``, ``float``, or :func:`_to_number`, which keeps int-vs-float) is applied
    otherwise; on failure the raw string is kept and the field is flagged
    ``needs_confirmation`` (never crash, never guess).
    """
    if py_type is str:
        return raw, "verified"
    if raw is None:
        return raw, "needs_confirmation"
    try:
        return py_type(raw), "verified"
    except (ValueError, TypeError):
        return raw, "needs_confirmation"


def _read_csv(path: Path) -> list[_Row]:
    """Read a ``section,field,value,unit,notes`` CSV into normalized rows."""
    rows: list[_Row] = []
    with open(path, newline="", encoding="utf-8") as fh:
        for record in csv.DictReader(fh):
            field = (record.get("field") or "").strip()
            if not field:
                continue
            section = (record.get("section") or "").strip()
            value = record.get("value")
            unit_raw = (record.get("unit") or "").strip()
            unit = unit_raw or None
            locator = f"Sheet '{section}', field={field}"
            rows.append(_Row(section, field, value, unit, locator))
    return rows


def _read_xlsx(path: Path) -> list[_Row]:
    """Read a multi-tab workbook (``field | value | unit | notes``) into rows.

    Each sheet's tab name is the ``section``; the value lives in column B, so the
    locator carries the real ``B<row>`` cell address of the value cell.
    """
    from openpyxl import load_workbook  # local import: optional dependency

    wb = load_workbook(filename=str(path), data_only=True)
    try:
        rows: list[_Row] = []
        for ws in wb.worksheets:
            section = ws.title
            # Row 1 is the header (field/value/unit/notes); data starts at row 2.
            for cells in ws.iter_rows(min_row=2):
                if len(cells) < 2:
                    continue
                field_cell, value_cell = cells[0], cells[1]
                unit_cell = cells[2] if len(cells) > 2 else None
                field = field_cell.value
                if field is None or str(field).strip() == "":
                    continue
                field = str(field).strip()
                raw_value = value_cell.value
                value = None if raw_value is None else str(raw_value)
                unit_val = unit_cell.value if unit_cell is not None else None
                unit = None if unit_val is None or str(unit_val).strip() == "" else str(unit_val).strip()
                locator = f"Sheet '{section}', cell {value_cell.coordinate}"
                rows.append(_Row(section, field, value, unit, locator))
    finally:
        wb.close()
    return rows


def _read_rows(path: Path) -> list[_Row]:
    """Dispatch to the CSV or XLSX reader on the file suffix."""
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError(f"unsupported structured artifact suffix: {path.suffix!r}")


def parse_structured(path) -> list[ExtractedField]:
    """Parse the campaign sheet into mapped scalar :class:`ExtractedField` list.

    Dispatches on suffix (``.xlsx`` via openpyxl, ``.csv`` via the csv module),
    normalizes to ``(section, field, value, unit)`` rows, then maps each row whose
    ``field`` is in :data:`FIELD_MAP`. Experimenters and unmapped rows are skipped
    (see :func:`parse_contributors`). Every emitted field is ``verified`` (or
    ``needs_confirmation`` on a failed numeric coercion) and carries exactly one
    ``spreadsheet`` evidence entry.
    """
    path = Path(path)
    source_file = path.name
    fields: list[ExtractedField] = []
    for row in _read_rows(path):
        mapping = FIELD_MAP.get(row.field)
        if mapping is None:
            continue  # contributor or 3B-deferred (series/qc) row: skip, never guess
        official_path, py_type = mapping
        value, status = _coerce(row.value, py_type)
        fields.append(
            ExtractedField(
                path=official_path,
                value=value,
                status=status,
                unit=row.unit,
                evidence=(_evidence(source_file, row.locator, row.value),),
            )
        )
    return fields


def parse_contributors(path) -> list[dict]:
    """Parse experimenter rows into contributor dicts for ``attribution`` (3B).

    Returns ``[{"name": ..., "role": "curated_record", "evidence": [...]}]``. These
    are deliberately kept out of :class:`ExtractedField` / ``fields`` — the Slice-3B
    builder decides how they populate ``attribution.contributors[]``.
    """
    path = Path(path)
    source_file = path.name
    contributors: list[dict] = []
    for row in _read_rows(path):
        if row.field in CONTRIBUTOR_FIELDS and row.value:
            contributors.append(
                {
                    "name": row.value,
                    "role": "curated_record",
                    "evidence": [_evidence(source_file, row.locator, row.value)],
                }
            )
    return contributors


def parse_rows(path) -> list[dict]:
    """Public accessor for the raw normalized sheet rows, INCLUDING the ones
    :func:`parse_structured` deliberately omits from scalar ``fields`` (``qc_status``
    and the incident-energy window belong to ``measurement`` blocks, not to a scalar
    path). The Slice-3B draft builder reads these to fill the ``qc`` block and the
    edge derivation note — it never guesses either. Each dict is
    ``{section, field, value, unit, locator, source_file}``.
    """
    path = Path(path)
    source_file = path.name
    return [
        {
            "section": r.section,
            "field": r.field,
            "value": r.value,
            "unit": r.unit,
            "locator": r.locator,
            "source_file": source_file,
        }
        for r in _read_rows(path)
    ]


__all__ = ["FIELD_MAP", "parse_structured", "parse_contributors", "parse_rows"]
